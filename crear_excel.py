import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Colores ──────────────────────────────────────────────────────────────────
AZUL       = "1F4E79"
AZUL_CLARO = "BDD7EE"
VERDE      = "375623"
VERDE_CLARO= "E2EFDA"
GRIS       = "F2F2F2"
AMARILLO   = "FFF2CC"
ROJO_CLARO = "FCE4D6"
BLANCO     = "FFFFFF"

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font_white_bold():
    return Font(color="FFFFFF", bold=True, size=10)

def font_bold():
    return Font(bold=True, size=10)

def font_normal():
    return Font(size=10)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row, cols_labels, fill_color):
    for col, label in enumerate(cols_labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = header_fill(fill_color)
        c.font = font_white_bold()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()

def style_data_row(ws, row, ncols, fill_hex=None):
    for col in range(1, ncols+1):
        c = ws.cell(row=row, column=col)
        if fill_hex:
            c.fill = PatternFill("solid", fgColor=fill_hex)
        c.font = font_normal()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()


# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — CARIES
# ════════════════════════════════════════════════════════════════════════════
ws_c = wb.active
ws_c.title = "CARIES"
ws_c.sheet_view.showGridLines = False

# Título
ws_c.merge_cells("A1:M1")
t = ws_c["A1"]
t.value = "PLANILLA REGISTRO — CARIES (ICDAS)  |  Estudio de Validación DEXA Vision"
t.fill = header_fill(AZUL)
t.font = Font(color="FFFFFF", bold=True, size=13)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[1].height = 30

# Subtítulo
ws_c.merge_cells("A2:M2")
s = ws_c["A2"]
s.value = "Unidad de análisis: imagen / cuadrante    |    ICDAS: 0 = Sano  ·  1-2 = Caries inicial  ·  3-4 = Caries moderada  ·  5-6 = Caries severa"
s.fill = header_fill(AZUL_CLARO)
s.font = Font(color="1F4E79", bold=False, size=9, italic=True)
s.alignment = Alignment(horizontal="center", vertical="center")
ws_c.row_dimensions[2].height = 18

# Cabeceras por grupo
ws_c.merge_cells("A3:D3")
ws_c["A3"].value = "IDENTIFICACIÓN"
ws_c["A3"].fill = header_fill("2E75B6")
ws_c["A3"].font = font_white_bold()
ws_c["A3"].alignment = Alignment(horizontal="center")

ws_c.merge_cells("E3:G3")
ws_c["E3"].value = "EVALUACIÓN CLÍNICA (EXAMINADORES)"
ws_c["E3"].fill = header_fill("375623")
ws_c["E3"].font = font_white_bold()
ws_c["E3"].alignment = Alignment(horizontal="center")

ws_c.merge_cells("H3:J3")
ws_c["H3"].value = "MODELO DEXA"
ws_c["H3"].fill = header_fill("7030A0")
ws_c["H3"].font = font_white_bold()
ws_c["H3"].alignment = Alignment(horizontal="center")

ws_c.merge_cells("K3:M3")
ws_c["K3"].value = "CONTROL CALIDAD"
ws_c["K3"].fill = header_fill("C55A11")
ws_c["K3"].font = font_white_bold()
ws_c["K3"].alignment = Alignment(horizontal="center")

ws_c.row_dimensions[3].height = 20

# Cabeceras columnas
headers_c = [
    "ID Imagen","Tipo Foto","Cuadrante","Calidad Imagen",
    "Exam1\nICDAS","Exam2\nICDAS","Referencia\nICDAS",
    "DEXA\nICDAS","Score\nDEXA (0-1)","Resultado\nDEXA",
    "Concord.\nExaminadores","Error\nDEXA","Observaciones"
]
apply_header(ws_c, 4, headers_c, "404040")
ws_c.row_dimensions[4].height = 35

# Datos de ejemplo
ejemplos_c = [
    ["IMG_001","Cuadrante","Sup. Derecho","Buena",0,0,0,0,0.04,"Sano","Sí","No",""],
    ["IMG_002","Cuadrante","Sup. Izquierdo","Buena",4,4,4,4,0.88,"ICDAS 4","Sí","No",""],
    ["IMG_003","Cuadrante","Inf. Derecho","Buena",2,3,2,2,0.61,"ICDAS 2","No — consenso Exam1","No","Discrepancia leve"],
    ["IMG_004","Arco completo","Ambos","Buena",5,5,5,4,0.91,"ICDAS 4","Sí","Sí — diferencia 1","Score alto, clasif. -1"],
    ["IMG_005","Cuadrante","Inf. Izquierdo","Regular",1,1,1,1,0.33,"ICDAS 1","Sí","No","Imagen con sombra"],
    ["IMG_006","Cuadrante","Sup. Derecho","No evaluable","","","","","","","","","Desenfocada — excluida"],
]

for i, row_data in enumerate(ejemplos_c):
    r = i + 5
    fill = GRIS if i % 2 == 0 else BLANCO
    if row_data[3] == "No evaluable":
        fill = ROJO_CLARO
    for col, val in enumerate(row_data, 1):
        c = ws_c.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = font_normal()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws_c.row_dimensions[r].height = 18

# Filas vacías para completar (hasta IMG_100)
for i in range(len(ejemplos_c), 100):
    r = i + 5
    fill = GRIS if i % 2 == 0 else BLANCO
    img_id = f"IMG_{str(i+1).zfill(3)}"
    ws_c.cell(row=r, column=1, value=img_id).fill = PatternFill("solid", fgColor=fill)
    for col in range(1, 14):
        c = ws_c.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = border_thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_c.row_dimensions[r].height = 16

# Anchos columnas caries
col_widths_c = [10, 13, 16, 14, 9, 9, 11, 9, 12, 12, 18, 15, 20]
for i, w in enumerate(col_widths_c, 1):
    ws_c.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — GINGIVITIS
# ════════════════════════════════════════════════════════════════════════════
ws_g = wb.create_sheet("GINGIVITIS")
ws_g.sheet_view.showGridLines = False

ws_g.merge_cells("A1:R1")
t2 = ws_g["A1"]
t2.value = "PLANILLA REGISTRO — GINGIVITIS  |  Estudio de Validación DEXA Vision"
t2.fill = header_fill(VERDE)
t2.font = Font(color="FFFFFF", bold=True, size=13)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws_g.row_dimensions[1].height = 30

ws_g.merge_cells("A2:R2")
s2 = ws_g["A2"]
s2.value = "Severidad: Sano = sin signos  ·  Leve = 1-2 signos leves  ·  Moderada = signos claros  ·  Severa = signos múltiples intensos"
s2.fill = header_fill(VERDE_CLARO)
s2.font = Font(color="375623", bold=False, size=9, italic=True)
s2.alignment = Alignment(horizontal="center", vertical="center")
ws_g.row_dimensions[2].height = 18

ws_g.merge_cells("A3:D3")
ws_g["A3"].value = "IDENTIFICACIÓN"
ws_g["A3"].fill = header_fill("2E75B6")
ws_g["A3"].font = font_white_bold()
ws_g["A3"].alignment = Alignment(horizontal="center")

ws_g.merge_cells("E3:I3")
ws_g["E3"].value = "SIGNOS CLÍNICOS VISIBLES (marcar Sí/No)"
ws_g["E3"].fill = header_fill("375623")
ws_g["E3"].font = font_white_bold()
ws_g["E3"].alignment = Alignment(horizontal="center")

ws_g.merge_cells("J3:L3")
ws_g["J3"].value = "EVALUACIÓN CLÍNICA"
ws_g["J3"].fill = header_fill("375623")
ws_g["J3"].font = font_white_bold()
ws_g["J3"].alignment = Alignment(horizontal="center")

ws_g.merge_cells("M3:O3")
ws_g["M3"].value = "MODELO DEXA"
ws_g["M3"].fill = header_fill("7030A0")
ws_g["M3"].font = font_white_bold()
ws_g["M3"].alignment = Alignment(horizontal="center")

ws_g.merge_cells("P3:R3")
ws_g["P3"].value = "CONTROL CALIDAD"
ws_g["P3"].fill = header_fill("C55A11")
ws_g["P3"].font = font_white_bold()
ws_g["P3"].alignment = Alignment(horizontal="center")

ws_g.row_dimensions[3].height = 20

headers_g = [
    "ID Imagen","Tipo Foto","Cuadrante","Calidad Imagen",
    "Inflamación","Eritema","Recesión","Cálculo\nVisible","Sangrado\nVisible",
    "Exam1\nSeveridad","Exam2\nSeveridad","Referencia\nSeveridad",
    "DEXA\nSeveridad","Score\nDEXA (0-1)","Resultado\nDEXA",
    "Concord.\nExaminadores","Error\nDEXA","Observaciones"
]
apply_header(ws_g, 4, headers_g, "404040")
ws_g.row_dimensions[4].height = 35

ejemplos_g = [
    ["IMG_001","Cuadrante","Sup. Derecho","Buena","No","No","No","No","No","Sano","Sano","Sano","Sano",0.06,"Sano","Sí","No",""],
    ["IMG_002","Cuadrante","Sup. Izquierdo","Buena","Sí","Sí","No","No","Sí","Leve","Leve","Leve","Leve",0.44,"Leve","Sí","No",""],
    ["IMG_003","Cuadrante","Inf. Derecho","Buena","Sí","Sí","Sí","Sí","Sí","Severa","Moderada","Severa","Moderada",0.79,"Moderada","No — consenso Exam1","Sí","Falso negativo parcial"],
    ["IMG_004","Arco completo","Ambos","Buena","Sí","Sí","No","Sí","No","Moderada","Moderada","Moderada","Moderada",0.71,"Moderada","Sí","No",""],
    ["IMG_005","Cuadrante","Inf. Izquierdo","Regular","No","Sí","No","No","No","Leve","Leve","Leve","Sano",0.28,"Sano","Sí","Sí — falso neg.","Score bajo"],
    ["IMG_006","Cuadrante","Sup. Derecho","No evaluable","","","","","","","","","","","","","","Desenfocada — excluida"],
]

for i, row_data in enumerate(ejemplos_g):
    r = i + 5
    fill = GRIS if i % 2 == 0 else BLANCO
    if row_data[3] == "No evaluable":
        fill = ROJO_CLARO
    for col, val in enumerate(row_data, 1):
        c = ws_g.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = font_normal()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws_g.row_dimensions[r].height = 18

for i in range(len(ejemplos_g), 100):
    r = i + 5
    fill = GRIS if i % 2 == 0 else BLANCO
    img_id = f"IMG_{str(i+1).zfill(3)}"
    ws_g.cell(row=r, column=1, value=img_id)
    for col in range(1, 19):
        c = ws_g.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = border_thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_g.row_dimensions[r].height = 16

col_widths_g = [10,13,16,14,12,10,10,10,10,12,12,13,13,12,12,18,15,20]
for i, w in enumerate(col_widths_g, 1):
    ws_g.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# HOJA 3 — INSTRUCCIONES
# ════════════════════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("INSTRUCCIONES")
ws_i.sheet_view.showGridLines = False

ws_i.column_dimensions["A"].width = 22
ws_i.column_dimensions["B"].width = 70

instrucciones = [
    ("INSTRUCCIONES DE USO", None, AZUL, True, 14),
    ("", None, BLANCO, False, 11),
    ("FLUJO DE TRABAJO", None, "2E75B6", True, 11),
    ("Paso 1", "Selecciona 100 imágenes de la base de datos. Nómbralas IMG_001 a IMG_100.", AZUL_CLARO, False, 10),
    ("Paso 2", "Procesa cada imagen con el modelo DEXA. Anota ICDAS, severidad de gingivitis y score en las planillas.", BLANCO, False, 10),
    ("Paso 3", "Examinador 1 evalúa todas las imágenes de forma independiente. Anota sus resultados.", AZUL_CLARO, False, 10),
    ("Paso 4", "Examinador 2 evalúa todas las imágenes (en sesión separada, sin ver resultados de Exam1 ni DEXA).", BLANCO, False, 10),
    ("Paso 5", "Completa la columna 'Referencia': en caso de acuerdo entre examinadores, usa ese valor. En caso de discrepancia, usa Exam1 y anota en Observaciones.", AZUL_CLARO, False, 10),
    ("Paso 6", "Envía el archivo Excel completo para el análisis estadístico.", BLANCO, False, 10),
    ("", None, BLANCO, False, 10),
    ("ESCALA ICDAS", None, VERDE, True, 11),
    ("ICDAS 0", "Sano — sin evidencia de caries.", VERDE_CLARO, False, 10),
    ("ICDAS 1", "Cambio blanco/marrón visible en esmalte seco.", BLANCO, False, 10),
    ("ICDAS 2", "Cambio blanco/marrón visible en esmalte húmedo.", VERDE_CLARO, False, 10),
    ("ICDAS 3", "Ruptura localizada del esmalte sin dentina visible.", BLANCO, False, 10),
    ("ICDAS 4", "Sombra oscura de dentina visible a través del esmalte.", VERDE_CLARO, False, 10),
    ("ICDAS 5", "Cavidad visible con dentina expuesta (<50% superficie).", BLANCO, False, 10),
    ("ICDAS 6", "Cavidad extensa con dentina expuesta (≥50% superficie).", VERDE_CLARO, False, 10),
    ("", None, BLANCO, False, 10),
    ("ESCALA GINGIVITIS", None, VERDE, True, 11),
    ("Sano", "Sin signos de inflamación gingival visibles en la fotografía.", VERDE_CLARO, False, 10),
    ("Leve", "1–2 signos presentes de forma leve: eritema leve o inflamación discreta.", BLANCO, False, 10),
    ("Moderada", "2–3 signos claros: eritema, inflamación, cálculo o sangrado visible.", VERDE_CLARO, False, 10),
    ("Severa", "Múltiples signos intensos: eritema marcado, recesión, cálculo abundante, sangrado evidente.", BLANCO, False, 10),
    ("", None, BLANCO, False, 10),
    ("CALIDAD DE IMAGEN", None, "C55A11", True, 11),
    ("Buena", "Imagen nítida, bien iluminada, encuadre correcto del cuadrante.", AMARILLO, False, 10),
    ("Regular", "Imagen evaluable con limitaciones menores (sombra, ángulo leve).", BLANCO, False, 10),
    ("No evaluable", "Imagen desenfocada, sobreexpuesta o con obstrucción. Se excluye del análisis.", ROJO_CLARO, False, 10),
]

for row_i, (col_a, col_b, fill, bold, size) in enumerate(instrucciones, 1):
    ws_i.row_dimensions[row_i].height = 20
    ca = ws_i.cell(row=row_i, column=1, value=col_a)
    ca.fill = PatternFill("solid", fgColor=fill)
    ca.font = Font(bold=bold, size=size, color="FFFFFF" if fill in [AZUL, VERDE, "2E75B6", "C55A11"] else "000000")
    ca.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ca.border = border_thin()

    if col_b is not None:
        cb = ws_i.cell(row=row_i, column=2, value=col_b)
        cb.fill = PatternFill("solid", fgColor=fill)
        cb.font = Font(size=size, color="FFFFFF" if fill in [AZUL, VERDE, "2E75B6", "C55A11"] else "000000")
        cb.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        cb.border = border_thin()
    else:
        ws_i.merge_cells(f"A{row_i}:B{row_i}")

wb.save("/home/user/DEXA-landing/DEXA-Estudio-Planillas.xlsx")
print("Excel creado correctamente.")
